"""Transcribes the source workbook into the committed ``data/*.json`` files.

The workbook is loaded twice: ``data_only=False`` for formula strings and
data-validation lists (weapon types, the C80 achievement selector), and
``data_only=True`` for resolved static values (the ``_OptionsCrafter`` trait
texts are ``="..."`` cells, so the cached strings are needed).  Nothing here
is invented; it only transcribes the source.
"""

from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from forge_calculator.parse import parse_ore_power, parse_trait
from .formula_parser import STAT_CELLS, parse_stat_matrix

DPS_SHEET = "DPS Calculator"
WEAPON_SHEET = "_Weapon"
CRAFTER_SHEET = "_OptionsCrafter"
OPTIONS_SHEET = "_Options"


# --- small helpers ---


def _in_range(sqref, coord):
    """True if the single-cell coord falls inside the (possibly multi-range) sqref."""
    try:
        min_col, min_row, max_col, max_row = range_boundaries(sqref)
    except Exception:
        return coord in str(sqref)
    r, c = coordinate_to_tuple(coord)  # note: (row, column)
    return min_row <= r <= max_row and min_col <= c <= max_col


def _find_validation(ws, target):
    """Return the data-validation formula1 string covering ``target``, or None."""
    dvs = getattr(getattr(ws, "data_validations", None), "dataValidation", []) or []
    for dv in dvs:
        if _in_range(str(dv.sqref), target):
            return dv.formula1
    return None


def _eval_excel_string_concat(s):
    """Resolve Excel ``"a,b"&"c,d"`` concatenations into a single string."""
    parts = []
    buf = []
    in_str = False
    for c in s or "":
        if c == '"':
            in_str = not in_str
        elif c == "&" and not in_str:
            parts.append("".join(buf))
            buf = []
        else:
            buf.append(c)
    parts.append("".join(buf))
    return "".join(parts)


def _validation_list(ws, target):
    """Return the parsed item list of an inline data-validation covering target."""
    f1 = _find_validation(ws, target)
    if not f1:
        return None
    text = _eval_excel_string_concat(f1)
    items = [p.strip() for p in text.split(",") if p.strip()]
    return items or None


def _cell_text(ws_v, ws_f, coord):
    """Best-effort text of a cell: cached value, falling back to ``="..."``."""
    v = ws_v[coord]
    val = v.value
    if val is not None:
        return val
    raw = ws_f[coord].value
    if isinstance(raw, str) and raw.startswith("="):
        raw = raw[1:].strip()
        if len(raw) >= 2 and raw.startswith('"') and raw.endswith('"'):
            return raw[1:-1]
    return raw


# --- extraction sections ---


def extract_weapons(ws_v):
    """_Weapon!B2:E80 -> (weapons, types)."""
    weapons = []
    types_ordered = []
    for row in ws_v.iter_rows(min_row=2, max_row=80):
        wtype, name, interval, damage = row[1].value, row[2].value, row[3].value, row[4].value
        name = (name or "").strip()
        if not name:
            continue
        weapons.append(
            {
                "name": name,
                "type": (wtype or "").strip(),
                "interval": float(interval),
                "damage": float(damage),
            }
        )
        if (wtype or "").strip() and (wtype or "").strip() not in types_ordered:
            types_ordered.append((wtype or "").strip())
    return weapons, types_ordered


def extract_ores(ws_v):
    """_OptionsCrafter!A56:F195 -> ore records (skips the A55 spacer)."""
    ores = []
    for row in ws_v.iter_rows(min_row=56, max_row=195):
        name = (row[0].value or "").strip()
        if not name:
            continue
        multiplier = parse_ore_power(row[2].value)
        ores.append(
            {
                "name": name,
                "equipment": (row[1].value or "").strip() or None,
                "multiplier": multiplier,
                "trait10": (row[3].value or "").strip() or None,
                "trait30": (row[4].value or "").strip() or None,
                "comments": (row[5].value or "").strip() or None,
            }
        )
    return ores


def extract_races(ws_v):
    """_Options race list + the availability matrix.

    Layout quirk in the workbook: the matrix rows ``E2:U17`` are indexed by the
    race names in column E (rows 2..17), while the dropdown list lives in
    ``B1:B16`` with descriptions in ``C1:C16`` (off by one row).  We read the
    names from column E (self-consistent with the matrix) and pull the default
    trait from the B->C map.
    """
    headers = []
    for col in range(6, 21):  # F..U
        headers.append((ws_v.cell(row=1, column=col).value or "").strip())
    # B1:B16 -> C1:C16 default-trait map (B[r] == E[r+1])
    default_traits = {}
    for r in range(1, 17):
        name = (ws_v.cell(row=r, column=2).value or "").strip()
        if name:
            default_traits[name] = (ws_v.cell(row=r, column=3).value or "").strip()
    races = []
    for r in range(2, 18):  # rows 2..17 == the 16 races
        name = (ws_v.cell(row=r, column=5).value or "").strip()
        if not name:
            continue
        avail = []
        for idx, header in enumerate(headers):
            cell = ws_v.cell(row=r, column=6 + idx).value
            if str(cell).strip().lower() in ("true", "1", "yes", "x"):
                avail.append(header)
        races.append(
            {
                "name": name,
                "default_trait": default_traits.get(name) or None,
                "available_traits": avail,
            }
        )
    return races, headers


def extract_runes(ws_v):
    """_Options!A1:A48 -> the rune trait pool (classified by parse_trait).

    The ``None`` placeholder is recorded separately in the ``none`` field and
    excluded from the pool list (47 traits).
    """
    runes = []
    for r in range(1, 49):
        name = (ws_v.cell(row=r, column=1).value or "").strip()
        if not name:
            continue
        if name == "None":
            continue
        parsed = parse_trait(name)
        if parsed:
            stat, value = parsed
            runes.append({"name": name, "stat": stat, "value": value})
        else:
            runes.append({"name": name})
    return runes


def extract_achievements(ws_f):
    """C80 data-validation list -> the 16 achievement-selector options."""
    names = _validation_list(ws_f[DPS_SHEET], "C80") or []
    achievements = []
    for name in names:
        parsed = parse_trait(name)
        if parsed:
            stat, value = parsed
            achievements.append({"name": name, "stat": stat, "value": value})
        else:
            achievements.append({"name": name})
    return achievements


def _signature(formulas):
    parts = sorted(
        f"{cell}={formulas.get(cell, '')}" for cell in STAT_CELLS
    )
    return hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()


# --- main entry ---


def extract(workbook_path, out_dir, source_name=None):
    wb_f = load_workbook(workbook_path, data_only=False)
    wb_v = load_workbook(workbook_path, data_only=True)

    weapons, derived_types = extract_weapons(wb_v[WEAPON_SHEET])
    ores = extract_ores(wb_v[CRAFTER_SHEET])
    races, trait_headers = extract_races(wb_v[OPTIONS_SHEET])
    runes = extract_runes(wb_v[OPTIONS_SHEET])
    achievements = extract_achievements(wb_f)

    # formula strings from the workbook (authoritative stat matrix)
    formulas = {}
    dps = wb_f[DPS_SHEET]
    for cell in STAT_CELLS:
        formulas[cell] = dps[cell].value or ""
    matrix = parse_stat_matrix(formulas)

    # weapon-type validation lists (C12 / C23) -- preserved verbatim
    types_from_c12 = _validation_list(dps, "C12")
    race_bonus_types = _validation_list(dps, "C23")
    if not types_from_c12:
        types_from_c12 = derived_types
    if not race_bonus_types:
        race_bonus_types = derived_types

    # attach stat matrix to catalog ores (by exact name)
    for ore in ores:
        ore["stats"] = matrix.get(ore["name"], {})

    # ores that appear only in formulas (not the catalog) -- preserved as a note
    formula_only = sorted(set(matrix) - {o["name"] for o in ores})

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    (out / "weapons.json").write_text(
        json.dumps(
            {
                "source": f"{WEAPON_SHEET}!C2:E80 (+ DPS C12/C23 validation lists)",
                "types": types_from_c12,
                "race_bonus_types": race_bonus_types,
                "weapons": weapons,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (out / "ores.json").write_text(
        json.dumps(
            {
                "source": f"{CRAFTER_SHEET}!A56:F195 + formulas C44..C75",
                "select_ore": "Select Ore",
                "formula_only_ores": formula_only,
                "ores": ores,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (out / "races.json").write_text(
        json.dumps(
            {
                "source": f"{OPTIONS_SHEET}!B1:C16 + availability matrix E2:U17",
                "trait_headers": trait_headers,
                "races": races,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (out / "runes.json").write_text(
        json.dumps(
            {
                "source": f"{OPTIONS_SHEET}!A1:A48",
                "none": "None",
                "runes": runes,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (out / "achievements.json").write_text(
        json.dumps(
            {
                "source": "DPS Calculator C80 data-validation list",
                "achievements": achievements,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (out / "meta.json").write_text(
        json.dumps(
            {
                "workbook": source_name or Path(workbook_path).name,
                "built_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                "counts": {
                    "ores": len(ores),
                    "weapons": len(weapons),
                    "races": len(races),
                    "runes": len(runes),
                    "achievements": len(achievements),
                },
                "stat_formulas_sha256": _signature(formulas),
                "formula_only_ores": formula_only,
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return {
        "counts": {
            "ores": len(ores),
            "weapons": len(weapons),
            "races": len(races),
            "runes": len(runes),
            "achievements": len(achievements),
        },
        "types": types_from_c12,
        "race_bonus_types": race_bonus_types,
        "formula_only_ores": formula_only,
    }
